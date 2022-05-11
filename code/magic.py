from ciscoconfparse import CiscoConfParse
import nmap3
import json
import configparser
from openpyxl import Workbook
from napalm import get_network_driver
import sys
import os
from types import SimpleNamespace
from pathlib import Path
from datetime import datetime
import ipaddress
import traceback
import concurrent.futures
from netmiko.ssh_autodetect import SSHDetect
from netmiko.ssh_dispatcher import ConnectHandler
import re
import logging


# TODO: Dump Device Data to json file for each device, generate tables using structured json data, move nmap


os.makedirs("./outputs/", exist_ok=True)

# TODO: If needed, open previous log file if exists and create append log

logging.basicConfig(filename='./outputs/log.txt', level=logging.WARNING, filemode='w', format='%(asctime)s %(levelname)-8s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')



#Def Functions
def DataFetchSafeHandler(dataFetchFunction, remoteDevice):
    try:
        return dataFetchFunction()

    except Exception as ex:
        swVersion = "Unknown"

        if "swVersion" in remoteDevice:
            swVersion = remoteDevice["swVersion"]

        errorMessage = f"Failed to fetch info for device with ip {remoteDevice['host']}, deviceType: {remoteDevice['device_type']} and software version: {swVersion}."
        print(errorMessage)
        logging.error(errorMessage)
        logging.error(traceback.format_exc())
        return None


def FetchNmapData(ipAddress):

    nmap = nmap3.Nmap()
    results = nmap.scan_top_ports(ipAddress)
    return results

def CreateNapalmConnection(ipAddress, driver, sshUserName, sshPassword):

    #NapalmConnection
    driver = get_network_driver(driver)
    device = driver(ipAddress, sshUserName, sshPassword)
    device.open()

    return device


def DeviceInfoFetchPipeline(ipAddress, username, password, timeout):

    swVersion = "unresolved"

    napalmDriverName = None

    remoteDevice = {
        'device_type': 'autodetect',
        'host': ipAddress,
        'username': username,
        'password': password }

    guesser = SSHDetect(**remoteDevice)
    bestMatch = guesser.autodetect() # TODO: rename bestMatch argument

    #print(bestMatch) # Name of the best device_type to use further
    #print(guesser.potential_matches) # Dictionary of the whole matching result

    if bestMatch == None:
        remoteDevice['device_type'] = "cisco_s300"
        connection = ConnectHandler(**remoteDevice)

        commandOutput = connection.send_command('show version', expect_string="#")

        regexVersionMatches = re.findall("\d+(?:\.\d+)+", commandOutput)

        if regexVersionMatches: napalmDriverName = 's350'

        else: raise Exception(f'Unrecognized Device. Tried to resolve device as Cisco SMB! Show version command output: {commandOutput}, regexVersionMatches: {regexVersionMatches}')

        remoteDevice['swVersion'] = regexVersionMatches[0]

    elif "huawei" in bestMatch:
        remoteDevice['device_type'] = bestMatch
        connection = ConnectHandler(**remoteDevice)
        
        commandOutput = connection.send_command('display version')

        if "VRP (R) software," in commandOutput:
            vrpVersion = str(commandOutput.partition("Version ")[2][0])

            if vrpVersion == "5": napalmDriverName = 'huawei_vrp'
            elif vrpVersion == "8": napalmDriverName = 'ce'
            else: raise Exception("Unrecognized Device. Tried to resolve device as Huawei Vrp5 & Vrp8!")

        if napalmDriverName == None: raise Exception("Unrecognized Device. Tried to resolve device as Huawei Vrp5 & Vrp8!")

        regexVersionMatches = re.findall("\d+(?:\.\d+)+.*", commandOutput)

        if regexVersionMatches: swVersion = regexVersionMatches[0]
        remoteDevice['swVersion'] = swVersion



    elif "cisco" in bestMatch: napalmDriverName = "ios"



    #TODO: Try default passwords
        


    #Fetch info through Napalm
    connection = CreateNapalmConnection(ipAddress, napalmDriverName, username, password)

    interfacesIp = DataFetchSafeHandler(lambda: ConvertDictInDictToDictInList(connection.get_interfaces_ip(), "interface"), remoteDevice)

    interfacesWithIp = []

    if (interfacesIp):
        for ipInterface in interfacesIp:
            ipsWithMasks = ''

            for ipAddressDictionary in ConvertDictInDictToDictInList(ipInterface['ipv4'], 'ip'):
                ipsWithMasks += ipAddressDictionary['ip'] + '/' + str(ipAddressDictionary['prefix_length']) + ", "

            if ipsWithMasks[-2:] == ', ':
                ipsWithMasks = ipsWithMasks[:-2]

            interfacesWithIp.append({ 
                'interface': ipInterface['interface'],
                'ipsWithMasks': ipsWithMasks})
    else:
        interfacesWithIp = None


    lldpNeighborsNapalmResult = DataFetchSafeHandler(lambda: connection.get_lldp_neighbors().items(), remoteDevice)

    lldpNeighbors = []

    if (lldpNeighborsNapalmResult):
        for key, value in lldpNeighborsNapalmResult:
            lldpNeighbors.append({
                "localInterface": key,
                "neighborInterface": value[0]["port"],
                "neigborHostname": value[0]["hostname"]})
    else:
        lldpNeighbors = None

    # compatibility matrix switch - YAMAN!

    napalmData = {
       "deviceConfig": DataFetchSafeHandler(lambda: str(connection.get_config()["running"]), remoteDevice),
       "arpTable" : DataFetchSafeHandler(lambda: connection.get_arp_table(), remoteDevice),
       "interfaces" : DataFetchSafeHandler(lambda: ConvertDictInDictToDictInList(connection.get_interfaces(), "interface"), remoteDevice),
       "interfacesIp" : interfacesWithIp,
       "lldpNeighbors" : lldpNeighbors}

    if not napalmDriverName == 's350':
        napalmData["macTable"] = DataFetchSafeHandler(lambda: connection.get_mac_address_table(), remoteDevice)
        napalmData["interfacesCounter"] = DataFetchSafeHandler(lambda: ConvertDictInDictToDictInList(connection.get_interfaces_counters(), "interfaceinterfacesCounter"), remoteDevice)
        napalmData["deviceUsers"] = DataFetchSafeHandler(lambda: connection.get_users(), remoteDevice)

    if napalmDriverName == "huawei_vrp":
        pass

    if napalmDriverName == "ce":
        pass

    if napalmDriverName == "ios":
        pass

    return {
        'ip': ipAddress,
        'napalmDriverName': napalmDriverName,
        'napalmData': napalmData,
        'nmapData': DataFetchSafeHandler(lambda: FetchNmapData(ipAddress), remoteDevice),
        'swVersion': swVersion
    }   

# TODO: Remove column feature for is_enabled in interfaces
# TODO: Ensure order of columns

def CreateWorksheet(workBook, workSheetName, dataCollection):

    if not dataCollection:
        #todo: log
        return

    workSheet = workBook.create_sheet(workSheetName)
    workSheet.append(list(dataCollection[0].keys()))

    for data in dataCollection:
        workSheet.append(list(data.values()))
    
def ConvertDictInDictToDictInList(dataDictDict, newColumnName):

    listDict = []


    for key, valueDict in dataDictDict.items():

        #Vytvarime novy dict s key=newColumnName a value=parentKey (GigabitEthernet0/0/3)
        dict = { newColumnName: key }

        #Appendneme zbytek key value hodnot za nove vytvoreny dict
        dict.update(valueDict)

        #novy dict hoduime do listu, ktery reprezenetuje 2d tabulku
        listDict.append(dict)


    return listDict


def SaveDeviceDataAsWorkbook(napalmData, path):
    
    workBook = Workbook()

    if "macTable" in napalmData:
        CreateWorksheet(workBook, "macTable", napalmData["macTable"])

    if "arpTable" in napalmData:
        CreateWorksheet(workBook, "arpTable", napalmData["arpTable"])

    if "interfaces" in napalmData:
        CreateWorksheet(workBook, "interfaces", napalmData["interfaces"])
    
    if "interfacesCounter" in napalmData:
        CreateWorksheet(workBook, "interfacesCounter", napalmData["interfacesCounter"])

    if "lldpNeighbors" in napalmData:
        CreateWorksheet(workBook, "lldpNeighbors", napalmData["lldpNeighbors"])

    if "interfacesIp" in napalmData:
        CreateWorksheet(workBook, "interfacesIp", napalmData["interfacesIp"])

    #delete default sheet
    del workBook['Sheet']

    workBook.save(path)


def SaveDeviceConfigFile(deviceConfig, path):
    deviceConfigFile = open(path, "w",encoding='utf8')
    deviceConfigFile.write(deviceConfig)
    deviceConfigFile.close()


    # else: raise Exception(f"Invalid configuration - operating system {vendor} is not supported.")

#FETCHING CONFIG and gather variables
config = configparser.ConfigParser()
config.read('./config.conf')

#Targets
ipAddresessFromConfig = config["Targets"]["IpAddresessToScan"]
if len(ipAddresessFromConfig) != 0:
    ipAddresessFromConfigList = ipAddresessFromConfig.split(",")
else:
    ipAddresessFromConfigList = []

networkFromConfig = config["Targets"]["NetworkToScan"]
UseJsonFileWithTargets = config["Targets"].getboolean("UseJsonFileWithTargets")

#Timeout for ssh connection
timeoutFromConfig = config["Targets"]["Timeout"]

#Credentials
sshUserName = config["Credentials"]["SshUserName"]
sshPassword = config["Credentials"]["SshPassword"]


#snmpCommunityName = config["Credentials"]["SnmpCommunityName"]
deviceConfigurationSave = config["Outputs"]["DeviceConfigurationSave"]
jsonNmapRaw = config["Outputs"]["JsonNmapRaw"]

vlanTables = config["Outputs"]["VlanTables"]
macTable = config["Outputs"]["MacTable"]
arpTables = config["Outputs"]["ArpTables"]

lldpPNeigbors = config["Outputs"]["LldpPNeigbors"]

#InterfaceOptions
interfaceTables = config["Outputs"]["InterfaceTables"]
interfaceCountersTables = config["Outputs"]["InterfaceCountersTables"]
interfacesIp = config["Outputs"]["InterfacesIp"]


#FETCHING CONFIG

if networkFromConfig:
    print(networkFromConfig)

print("LIST", bool(ipAddresessFromConfigList))
print("LIST LEN", len(ipAddresessFromConfigList))

if UseJsonFileWithTargets: pass #Import Json data to ipAddresessToScan

else:
    if ipAddresessFromConfigList and networkFromConfig: #
        exitMessage = "Wrong settings detected. Choose only one from 'NetworkToScan' or 'IpAddresessToScan' in the config, let the other one blank !!!"
        logging.error(exitMessage)
        sys.exit(exitMessage)
        
    elif (not ipAddresessFromConfigList) and networkFromConfig: ipAddresessToScan = [str(ip) for ip in ipaddress.IPv4Network(networkFromConfig)]
    elif ipAddresessFromConfigList and (not networkFromConfig): ipAddresessToScan = ipAddresessFromConfigList
    else: 
        exitMessage = "Wrong settings detected. You must specify the targets !!!"
        logging.error(exitMessage)
        sys.exit(exitMessage)

allDeviceData = []

#Async Block with 256 threads
with concurrent.futures.ThreadPoolExecutor(max_workers=256) as executor:

    future_to_ipAddressToScan = {executor.submit(DeviceInfoFetchPipeline, ipAddressToScan, sshUserName, sshPassword, int(timeoutFromConfig)): ipAddressToScan for ipAddressToScan in ipAddresessToScan}

    for future in concurrent.futures.as_completed(future_to_ipAddressToScan):

        ipAddressToScan = future_to_ipAddressToScan[future]

        try:
            deviceData = future.result()

            deviceOutputFolderPath = "./outputs/devices/" + deviceData["ip"].replace(".", "_")

            os.makedirs(deviceOutputFolderPath, exist_ok=True)

            if jsonNmapRaw:
                nmapRawJson = open(deviceOutputFolderPath + "/nmapRaw.json", "w", encoding='utf8')
                nmapRawJson.write(json.dumps(deviceData['nmapData'], indent=4))
                nmapRawJson.close()
                deviceData["nmap"] = jsonNmapRaw

            allDeviceData.append(deviceData) # can create custom json list file to reduce RAM usage

            SaveDeviceConfigFile(deviceData["napalmData"]["deviceConfig"], deviceOutputFolderPath + "/config.txt")
            SaveDeviceDataAsWorkbook(deviceData["napalmData"], deviceOutputFolderPath + "/deviceInfo.xlsx")

            

            
                   
        except Exception as ex:
            logging.error(f'{ipAddressToScan} generated an exception: {ex}')
            logging.error(traceback.format_exc())

out_file = open("./outputs/allDeviceData.json", "w") 
json.dump(allDeviceData, out_file)
out_file.close()

#ENDE
print("Capo ti tuti capi ende slus !!!")

#os.system("capo.mp3")



# useTestingData=False
# createTestingData = False

# if useTestingData:
#         json.loads(open("./testSwitchData.json",'r').read())

# if createTestingData:
#                 with open("./testSwitchData.json", 'w') as outfile:
#                     json.dump(deviceData, outfile)